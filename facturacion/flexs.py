from datetime import datetime
from flask import (
    Blueprint, g, render_template, request, session,send_file
)
from auth import auth
from informeErrores import informeErrores
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font



from database import database
from scriptGeneral import scriptGeneral
fb = Blueprint('facturacion', __name__, url_prefix='/')


@fb.route("/descargaresumen")
@auth.login_required
def descargaResumen():
    return send_file('liquidacion.xlsx')

@fb.route('/consulta_flexs')
@auth.login_required
def consultaFlexs():
    hoy = str(datetime.now())[0:10]
    return render_template("facturacion/tabla_viajes.html",
                        titulo="Facturacion", 
                        desde=hoy,
                        hasta=hoy,
                        clientes=scriptGeneral.consultar_clientes(database.connect_db()), 
                        tipo_facturacion="flex", 
                        auth = session.get("user_auth"))

@fb.route("/facturacion_flex", methods=["GET"])
@auth.login_required
def facturacionFlex():
    midb = database.connect_db()
    cursor = midb.cursor()
    cliente = request.args.get("cliente")
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    viajes = []
    suma = 0
    book = Workbook()
    sheet = book.active
    image = Image("static/logo_grande.jpeg")
    sheet.add_image(image, 'A1')

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 65
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet["A9"] = "Fecha"
    sheet["B9"] = "Numero de envío"
    sheet["C9"] = "Comprador"
    sheet["D9"] = "Direccion_Completa"
    sheet["E9"] = "Localidad"
    sheet["F9"] = "Precio"
    sheet["E3"] = desde
    sheet["E4"] = hasta
    for row in sheet['A9:H9']:
        for cell in row:
            cell.fill = PatternFill(fgColor='FF0000', fill_type='solid')
            cell.font = Font(color='FFFFFF')
    
    
    contador = 9
    sql = f"""select H.Fecha, H.Numero_envío,H.Direccion_Completa,H.Localidad,H.Precio,V.comprador,V.Cobrar,V.estado_envio from historial_estados as H inner join ViajesFlexs as V on V.Numero_envío = H.Numero_envío where vendedor(H.Vendedor) = '{cliente}' and H.Fecha between '{desde}' and '{hasta}' and H.estado_envio in ('En Camino','Levantada') order by Fecha desc"""
    cursor.execute(sql)
    sinprecio = 0
    for viajeTupla in cursor.fetchall():
        viaje = list(viajeTupla)
        contador += 1
        fecha = viaje[0]
        nenvio = viaje[1]
        direccionCompleta = viaje[2]
        localidad = viaje[3]
        precio = viaje[4]
        comprador = viaje[5]
        cobrar = viaje[6]
        estado = viaje[7]
        if viaje[7] != "Entregado":
            viaje[6] = "0"

        if(precio == None):
            sinprecio = sinprecio +1
            precio = "Sin precio"
        else:
            suma = suma + float(precio)
        sheet["A"+str(contador)] = fecha
        sheet["B"+str(contador)] = nenvio
        sheet["C"+str(contador)] = comprador
        sheet["D"+str(contador)] = direccionCompleta
        sheet["E"+str(contador)] = localidad
        sheet["F"+str(contador)] = precio
        if estado == "Entregado":
            sheet["G"+str(contador)] = cobrar
        else:
            sheet["G"+str(contador)] = "0"
        sheet["H"+str(contador)] = estado
        viajes.append(viaje)
    sheet["E1"] = cliente
    sheet["E2"] = f"cantidad: {contador-9}"
    sheet["E3"] = "Subtotal: "
    sheet["E4"] = "IVA: "
    sheet["F6"] = "Total: "
    sheet["F3"] = "=SUM(F10:F"+str(contador)+")"
    sheet["F4"] = "=F3 * 0.21"
    sheet["F6"] = "=SUM(E3:E4)"
    book.save("liquidacion.xlsx")
    cabeceras = ["Fecha","Numero de envío","Direccion Completa","Localidad","Precio","Comprador"]
    return render_template("facturacion/tabla_viajes.html",
                            cliente=cliente,
                            desde=desde,
                            hasta=hasta,
                            titulo="Facturacion", 
                            cabeceras = cabeceras,
                            tipo_facturacion="flex", 
                            viajes=viajes, 
                            total=f"${suma} y {sinprecio} viajes sin precio", 
                            clientes = scriptGeneral.consultar_clientes(midb), 
                            auth = session.get("user_auth")
                            )

@fb.route("/facturacion/borrar_repetido")
@auth.login_required
@auth.admin_required
def borrarRepetidos():
    midb = database.connect_db()
    cursor = midb.cursor()
    cursor.execute("select id, Numero_envío, estado_envio from historial_estados where Numero_envío in (select Numero_envío from historial_estados where estado_envio = 'En Camino' group by Numero_envío having count(Numero_envío) >1) and estado_envio = 'En Camino'")
    envios = {}
    for x in cursor.fetchall():
        if x[1] in envios.keys():
            envios[str(x[1])].append(x[2])
        else:
            envios[x[1]] = [x[2]]
    cantidad = 0
    enviosDuplicados = []
    for x in envios:
        if (len(envios[x])) > 1:
            enviosDuplicados.append(x)
            cantidad += 1
    print(f"{cantidad} a revisar antes de facturar")
    borrados = 0
    controlado = 0
    errores = []
    for y in enviosDuplicados:
        controlado += 1
        # midb = database.verificar_conexion(midb)
        sql = f"select id from historial_estados where estado_envio = 'En Camino' and Numero_envío = '{y}'"
        cursor.execute(sql)
        enCaminos = 0
        seGuarda = 99999999999999999999999999999999999999999999999999999999999999999999999
        for x in cursor.fetchall():
            if x[0] < seGuarda:
                seGuarda = x[0]
            else:
                sql = f"DELETE FROM historial_estados WHERE id = '{x[0]}';"
                borrados+=1
                try:
                    cursor.execute(sql)
                    midb.commit()
                except Exception as err:
                    errores.append(err)

    cursor = midb.cursor()
    cursor.execute("select id, Numero_envío, estado_envio from historial_estados where estado_envio = 'Entregado' and Fecha >= '2022-09-01' group By Numero_envío having count(Numero_envío)>1 ")
    envios = {}
    entregadosBorrados = 0
    mayor = 0
    for y in cursor.fetchall():
        cursor.execute(f"select id, Numero_envío, estado_envio from historial_estados where Numero_envío = '{y[1]}'")
        for x in cursor.fetchall():
            if x[0] > mayor:
                sql = f"delete from historial_estados where id = {mayor}"
                entregadosBorrados += 1
                mayor = x[0]
            else:
                sql = f"delete from historial_estados where id = {x[0]}"
                entregadosBorrados += 1
            print(sql)
            print(entregadosBorrados)
            try:
                cursor.execute(sql)
                midb.commit()
            except Exception as err:
                errores.add(err)
    fechaEnvio = []
    agregados = 0
    cursor.execute(f"""select H.Fecha,H.Numero_envío,V.Localidad,V.Vendedor,V.Direccion,V.Precio_Cliente
        from historial_estados as H join ViajesFlexs as V on H.Numero_envío = V.Numero_envío
            where lower(H.estado_envio) in ('entregado','no entregado','listo para salir (sectorizado)') 
            and not H.Numero_envío in 
                (select Numero_envío from historial_estados where estado_envio in ('En Camino') or motivo_noenvio in ('Cancelado'))
                """)
    for x in cursor.fetchall():
        datoEnvio = [x[0],x[1],x[2],x[3],x[4],x[5]]
        fechaEnvio.append(datoEnvio)
    for enCamino in fechaEnvio:
        sql = f"insert into historial_estados (Fecha, Numero_envío,estado_envio,Localidad,Vendedor,Direccion_completa,Precio) values('{enCamino[0]}','{enCamino[1]}','En Camino','{enCamino[2]}','{enCamino[3]}','{enCamino[4]}','{enCamino[5]}');"
        agregados += 1
        cursor = midb.cursor()
        cursor.execute(sql)
        midb.commit()
    cant_err = len(errores)
    if cant_err != 0:
        return f"Se borraron {borrados} En Camino, {entregadosBorrados} Entregados y se agregaron {agregados} En Camino faltantes.\nSe produjeron {cant_err} errores"
    else:
        return f"Se borraron {borrados} En Camino, {entregadosBorrados} Entregados y se agregaron {agregados} En Camino faltantes."