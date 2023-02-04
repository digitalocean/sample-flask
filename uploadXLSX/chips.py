from flask import Blueprint, render_template, request, session,redirect
from threading import Thread
from datetime import datetime
from auth import auth
from database import database
from informeErrores import informeErrores
import openpyxl
from logistica.Envio import Envio
from logistica.script import geolocalizarFaltantes
from scriptGeneral import scriptGeneral
from threading import Thread
from tareasProgramadas.tareasProgramadas import generarInforme
formatSim = Blueprint('formatSim', __name__, url_prefix='/')

@formatSim.route("/logistica/carga_recorridos", methods = ["GET","POST"])
@auth.login_required
def subir_exel_chips():
    ahora = (datetime.today())
    if request.method == "POST":
        midb = database.connect_db() 
        cursor = midb.cursor()
        cursor.execute("select sim from GSolutions")
        envios = {}
        for x in cursor.fetchall():
            envios[x[0]] = True
        archivo_xlsx = request.files["upload"]
        libro = openpyxl.load_workbook(archivo_xlsx)
        sheet_obj = libro.active 
        cant_columnas = 25
        contador = 0
        try:
            for x in range(cant_columnas):
                contador += 1
                cab = str(sheet_obj.cell(row = 1, column = contador).value).lower().replace(" ","")
                print(cab)
                if cab == "fecha":
                    col_fecha = contador
                elif cab == ("nombre"):
                    col_nombre = contador
                elif cab == ("apellido"):
                    col_apellido = contador
                elif cab in ("cliente","firstName(shipping)","comprador","quienrecibe"):
                    col_cliente = contador
                elif cab in ("direccion","address1&2(Shipping)","dirección"):
                    col_direccion = contador
                elif cab in ("número","numero","altura"):
                    col_altura = contador
                elif cab in ("localidad","city(shipping)","ciudad"):
                    col_localidad = contador
                elif cab == ("vendedor"):
                    col_vendedor = contador
                elif cab in ("cp","postcode(shipping)","códigopostal","codigopostal","códigopostal"):
                    col_cp = contador
                elif cab in ("telefono","phone(billing)","lineaaportar"):
                    col_telefono = contador
                elif cab in ("referencia","referencias","customernote","detalledirección","detalledireccion"):
                    col_referencia = contador
                elif cab in "torre/monoblock":
                    col_TorreMonoblock = contador
                elif cab in "piso":
                    col_Piso = contador
                elif cab in "departamento":
                    col_Departamento = contador
                elif cab in "manzana":
                    col_Manzana = contador
                elif cab in "barrio":
                    col_Barrio = contador
                elif cab in "casa/lote":
                    col_casaLote = contador
                elif cab in "entre_calles":
                    col_EntreCalles = contador

            print("\nasignacion completa\n")
        except Exception as cabezeras:
            informeErrores.reporte(cabezeras,"/carga_formms")
            print("\nError al asignar variables\n")
        n_row = 1
        omitido = 0
        flex_agregado = 0
        cantidad = range(sheet_obj.max_row)
        contadorCantidad = 0
        for fila in cantidad:
            contadorCantidad += 1
            if contadorCantidad > 200:
                break
        if contadorCantidad > 200:
            cantidad = range(0,200)
        viajes = []
        chars = '.,!-"#$%&/()=?¡¿:[]'
        for x in cantidad:
            n_row += 1
            fecha = str(ahora)[0:10]
            if "col_fecha" in locals():
                fecha = str(sheet_obj.cell(row = n_row, column = col_fecha).value)
            if "col_cliente" in locals():
                cliente = str(sheet_obj.cell(row = n_row, column = col_cliente).value)
            else:
                if "col_nombre" in locals():
                    cliente = str(sheet_obj.cell(row = n_row, column = col_nombre).value)
                if "col_apellido" in locals():
                    cliente = cliente + " " + str(sheet_obj.cell(row = n_row, column = col_apellido).value)
            if "col_telefono" in locals():    
                telefono = str(sheet_obj.cell(row = n_row, column = col_telefono).value).replace("+54","").replace(" ","")
                telefono = telefono.translate(str.maketrans('', '', chars))
                if telefono[0:2] == "15":
                    telefono = "11" + telefono[2:]
            else:
                telefono = None
            if "col_direccion" in locals(): 
                direccion = str(sheet_obj.cell(row = n_row, column = col_direccion).value)
                if "col_altura" in locals():
                    direccion = direccion + str(sheet_obj.cell(row = n_row, column = col_altura).value)
            else:
                direccion = ""
            if "/" in str(direccion):
                direccion = str(direccion.split("/"))[0]
            if "col_referencia" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_referencia).value)
                if valor == "None":
                    observacion = "None"
                else:
                    observacion = valor
            referencia = ""
            if "col_TorreMonoblock" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_TorreMonoblock).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | TorreMonoblock: " + valor
            if "col_Piso" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_Piso).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | Piso: " + valor
            if "col_Departamento" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_Departamento).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | Departamento: " + valor
            if "col_Manzana" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_Manzana).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | Manzana: " + valor
            if "col_Barrio" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_Barrio).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | Barrio: " + valor
            if "col_casaLote" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_casaLote).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | casaLote: " + valor
            if "col_EntreCalles" in locals():
                valor = str(sheet_obj.cell(row = n_row, column = col_EntreCalles).value)
                if valor != "None" and valor != '""' and valor != "":
                    referencia = referencia + " | entreCalles: " + valor                
            if "col_localidad" in locals(): 
                localidad = str(sheet_obj.cell(row = n_row, column = col_localidad).value)
            if "col_cp" in locals():
                cp = str(sheet_obj.cell(row = n_row, column = col_cp).value)
            else:
                cp = 0
            if cp == "":
                cp = 0
            if session.get("user_auth") == "Cliente":
                vendedor = session.get("user_id")
            else:
                vendedor = request.form.get("nombre_cliente")
            
            fecha = fecha[0:10].replace("/","-").replace("\\","")
            if direccion == "None" or localidad == "None":
                continue
            informar = False
            chars = '.,!"#$%&/()=?¡¿'
            nro_envio = None
            viaje = Envio(direccion,localidad,vendedor,nro_envio,cliente,telefono,referencia,cp,fecha,tipoEnvio=15,sku="sim",fromDB=True,observacion=observacion)
            if referencia == "":
                referencia = observacion
            resu = viaje.toDB()
            if resu:
                viajes.append([resu,cliente,direccion,localidad,telefono,referencia])
                flex_agregado += 1 
            else: 
                omitido+=1
                viajes.append([nro_envio,cliente,"","","","No registrado, el numero de envio ya existe!"])
        cabezeras = ["Numero de envío","Cliente","Direccion","Localidad","Telefono","Referencia","Monto","Producto"]
        Tg = Thread(target=geolocalizarFaltantes, args=(database.connect_db(),))
        Tg.start()
        if informar:
            t = Thread(target=generarInforme, args=(database.connect_db(),'descargas/informe.xlsx',vendedor))
            t.start()
        return render_template("CargaArchivo/data.html",
                                titulo="Carga", 
                                data=viajes,
                                titulo_columna=cabezeras,
                                agregados=flex_agregado, 
                                repetido=omitido, 
                                auth = session.get("user_auth"))
    else:
        return render_template("CargaArchivo/carga_archivo.html",
                                titulo="Carga", 
                                clientes=scriptGeneral.consultar_clientes(database.connect_db()), 
                                auth = session.get("user_auth"), 
                                url_post="logistica/carga_recorridos")