#!/usr/bin/env python
# -*- coding: utf-8 -*- 
# encoding: utf-8

from flask import Blueprint, render_template, request, session,redirect
from threading import Thread
from datetime import datetime
from auth import auth
from database import database
from informeErrores import informeErrores
import openpyxl
from logistica import Envio
from logistica.script import geolocalizarFaltantes
from scriptGeneral import scriptGeneral
from threading import Thread
import pandas as pd
def generarInforme(midb,ruta,vendedor):
    pd.read_sql(f"select * from ViajesFlexs where Vendedor = '{vendedor}' and Fecha > current_date();",midb).to_excel(ruta)
    scriptGeneral.enviar_correo(["njb.11@hotmail.com","acciaiomatiassebastian@gmail.com"],f"Envios cargados {vendedor}",'descargas/informe.xlsx',"Informe.xlsx","")

formms = Blueprint('formms', __name__, url_prefix='/')

@formms.route("/carga_formms", methods = ["GET","POST"])
@auth.login_required
def subir_exel_formms():
    ahora = (datetime.today())
    if request.method == "POST":
        midb = database.connect_db()
        cursor = midb.cursor()
        cursor.execute("select Numero_envío from ViajesFlexs")
        envios = {}
        for x in cursor.fetchall():
            envios[x[0]] = True
        archivo_xlsx = request.files["upload"]
        libro = openpyxl.load_workbook(archivo_xlsx)
        sheet_obj = libro.active 
        cant_columnas = 14
        contador = 0
        try:
            for x in range(cant_columnas):
                contador += 1
                cab = str(sheet_obj.cell(row = 1, column = contador).value)
                if cab.lower() == "fecha":
                    col_fecha = contador
                elif cab.lower() in ("numero de envio", "numero de envío", "nro de envio","order number"):
                    col_numero_envio = contador
                elif cab.lower() in ("cliente","first Name (shipping)","comprador","quien recibe"):
                    col_cliente = contador
                elif cab.lower() in ("direccion","address 1&2 (Shipping)"):
                    col_direccion = contador
                elif cab.lower() in ("localidad","city (shipping)"):
                    col_localidad = contador
                elif cab.lower() in ("vendedor"):
                    col_vendedor = contador
                elif cab.lower() in ("cp","postcode (shipping)","código postal","codigo postal"):
                    col_cp = contador
                elif cab.lower() in ("telefono","phone (billing)"):
                    col_telefono = contador
                elif cab.lower() in ("referencia","customer note","detalle dirección","detalle direccion"):
                    col_referencia = contador
                elif cab.lower() in ("cobrar","monto","precio"):
                    col_cobrar = contador
                elif cab.lower() in ("producto"):
                    col_producto = contador
            print("\nasignacion completa\n")
        except Exception as cabezeras:
            informeErrores.reporte(cabezeras,"/carga_formms")
            print("\nError al asignar variables\n")
        n_row = 1
        omitido = 0
        flex_agregado = 0
        cantidad = range(sheet_obj.max_row)
        contadorCantidad = 0
        for fila in cantidad:
            contadorCantidad += 1
            if contadorCantidad > 200:
                break
        if contadorCantidad > 200:
            cantidad = range(0,200)
        viajes = []
        for x in cantidad:
            n_row += 1
            if "col_numero_envio" in locals():
                nro_envio = str(sheet_obj.cell(row = n_row, column = col_numero_envio).value)
            else:
                nro_envio = ""
            fecha = str(ahora)[0:10]
            if "col_fecha" in locals():
                fecha = str(sheet_obj.cell(row = n_row, column = col_fecha).value)
            if "col_cliente" in locals():
                cliente = str(sheet_obj.cell(row = n_row, column = col_cliente).value)
            else:
                cliente = ""
            if "col_telefono" in locals():    
                telefono = str(sheet_obj.cell(row = n_row, column = col_telefono).value)
            else:
                telefono = None
            if "col_direccion" in locals(): 
                direccion = str(sheet_obj.cell(row = n_row, column = col_direccion).value)
            else:
                direccion = ""
            if "/" in str(direccion):
                direccion = str(direccion.split("/"))[0]
            if "col_referencia" in locals():
                referencia = str(sheet_obj.cell(row = n_row, column = col_referencia).value)
            else:
                referencia = ""
            if "col_localidad" in locals(): 
                localidad = str(sheet_obj.cell(row = n_row, column = col_localidad).value)
            if "col_cp" in locals():
                cp = str(sheet_obj.cell(row = n_row, column = col_cp).value)
            else:
                cp = 0
            if cp == "":
                cp = 0
            if "col_cobrar" in locals():
                cobrar = str(sheet_obj.cell(row= n_row,column=col_cobrar).value)
            else:
                cobrar = 0
            if "col_producto" in locals():
                producto = str(sheet_obj.cell(row= n_row,column=col_producto).value)
            else:
                producto = "" 
            if session.get("user_auth") == "Cliente":
                vendedor = session.get("user_id")
            elif session.get("user_auth") == "Zippin":
                vendedor = str(sheet_obj.cell(row = n_row, column = col_vendedor).value)
            else:
                vendedor = request.form.get("nombre_cliente")
            
            fecha = fecha[0:10].replace("/","-").replace("\\","")
            tipo_envio = 2
            if direccion == "None" or localidad == "None":
                continue
            informar = False
            if vendedor == "Quality Shop" or vendedor == "Armin" or vendedor == "Happe":
                informar = True
                tipo_envio = 13
                nro_envio = None
            else:
                chars = '.,!"#$%&/()=?¡¿'
                if str(nro_envio).translate(str.maketrans('', '', chars)) in str(envios.keys()):
                    viajes.append([nro_envio,cliente,"","","","No registrado, el numero de envio ya existe!"])
                    omitido += 1
                    continue
            if nro_envio == "None":
                nro_envio = None
            
            print(nro_envio)
            viaje = Envio.Envio(direccion,localidad,vendedor,nro_envio,cliente,telefono,referencia,cp,fecha,tipoEnvio=tipo_envio,cobrar=cobrar,sku=producto,fromDB=True)
            resu = viaje.toDB()
            if resu:
                viajes.append([resu,cliente,direccion,localidad,telefono,referencia,cobrar,producto])
                flex_agregado += 1 
            else: 
                omitido+=1
                viajes.append([nro_envio,cliente,"","","","No registrado, el numero de envio ya existe!"])
        cabezeras = ["Numero de envío","Cliente","Direccion","Localidad","Telefono","Referencia","Monto","Producto"]
        t = Thread(target=geolocalizarFaltantes,args=(database.connect_db(),))
        t.start()
        if informar:
            t = Thread(target=generarInforme, args=(database.connect_db(),'descargas/informe.xlsx',vendedor))
            t.start()
        return render_template("CargaArchivo/data.html",
                                titulo="Carga", 
                                data=viajes,
                                titulo_columna=cabezeras,
                                agregados=flex_agregado, 
                                repetido=omitido, 
                                auth = session.get("user_auth"))

    else:
        return render_template("CargaArchivo/carga_archivo.html",
                                titulo="Carga", 
                                clientes=scriptGeneral.consultar_clientes(database.connect_db()), 
                                auth = session.get("user_auth"), 
                                url_post="carga_formms")


