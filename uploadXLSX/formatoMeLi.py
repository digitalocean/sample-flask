#!/usr/bin/env python
# -*- coding: utf-8 -*- 
# encoding: utf-8

from flask import Blueprint, render_template, request, session
from auth import auth
from logistica import Envio
import openpyxl
from database import database
from scriptGeneral import scriptGeneral


upML = Blueprint('uploadML', __name__, url_prefix='/')

@upML.route("/upload", methods=["GET","POST"])
@auth.login_required
def upload():
    if request.method == "POST":
        if session.get("user_auth") == "Cliente":
            vendedor = session.get("user_id")
        else:
            vendedor = request.form.get("nombre_cliente")
        file = request.files["upload"]
        wb_obj = openpyxl.load_workbook(file) 
        sheet_obj = wb_obj.active 
        cant_filas = sheet_obj.max_row
        n_row = 1
        indice_row = sheet_obj.cell(row = n_row, column = 1).value
        while indice_row != "# de venta":
            n_row += 1
            indice_row = str(sheet_obj.cell(row = n_row, column = 1).value)
        col_ini = 1
        indice_col = sheet_obj.cell(row = n_row, column = col_ini).value
        while indice_col != "Comprador":
            col_ini += 1
            indice_col = sheet_obj.cell(row = n_row, column = col_ini).value
        flex = 0
        flex_agregado = 0
        repetido = 0
        lista_viajes = []
        nros_envios = []
        midb = database.connect_db()
        cursor = midb.cursor()
        cursor.execute("select Numero_envío from ViajesFlexs")
        for x in cursor:
            nros_envios.append(x[0])
        cant_viajes = 0
        for x in range(cant_filas - n_row):
            n_row += 1 
            forma_entrega = sheet_obj.cell(row = n_row, column = col_ini + 7).value
            print(forma_entrega)
            if str(type(forma_entrega)) == "<class 'NoneType'>" or str(type(forma_entrega)) == "<type 'NoneType'>":
                pass
            else:
                cant_viajes += 1
                if "Flex" in forma_entrega:
                    flex += 1
                    nro_seguimiento = sheet_obj.cell(row = n_row, column = col_ini + 11).value
                    if nro_seguimiento in nros_envios:
                        repetido += 1
                    else:
                        nros_envios.append(nro_seguimiento)
                        nro_venta = sheet_obj.cell(row = n_row, column = col_ini + 11).value
                        flex_agregado += 1
                        estado_envio = sheet_obj.cell(row = n_row, column = 3).value
                        comprador = sheet_obj.cell(row = n_row, column = col_ini).value
                        dni = sheet_obj.cell(row = n_row, column = col_ini + 1).value
                        domicilio = sheet_obj.cell(row = n_row, column = col_ini + 2).value
                        domicilio = domicilio.split("/")
                        direccion = domicilio[0]
                        if len(domicilio) > 1:
                            referencia = domicilio[1]
                        else:
                            referencia = ""
                        ciudad  = sheet_obj.cell(row = n_row, column = col_ini + 3).value
                        estado= sheet_obj.cell(row = n_row, column = col_ini + 4).value
                        direccion_completa = direccion + ", " + ciudad + ", Buenos Aires"
                        cp = sheet_obj.cell(row = n_row, column = col_ini + 5).value
                        pais = sheet_obj.cell(row = n_row, column = col_ini + 6).value
                        fecha_encamino = sheet_obj.cell(row = n_row, column = col_ini + 8).value
                        fecha_entregado = sheet_obj.cell(row = n_row, column = col_ini + 9).value
                        transportista = sheet_obj.cell(row = n_row, column = col_ini + 10).value
                        url_seguimiento = sheet_obj.cell(row = n_row, column = col_ini + 12).value
                        paquete = [estado_envio,comprador,dni,direccion,ciudad,estado,cp,referencia,pais,fecha_encamino,fecha_entregado,transportista,nro_seguimiento,url_seguimiento,forma_entrega]
                        viaje = Envio.Envio(direccion,ciudad,vendedor,nro_seguimiento,comprador,referencia=referencia,numeroVenta=nro_venta,cp=cp)
                        viaje.toDB()
                        # sql = "insert into ViajesFlexs (Fecha, Numero_envío,nro_venta,comprador,Direccion,Referencia,Localidad,CP,Vendedor,estado_envio, Direccion_completa,tipo_envio) values (current_date(),%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,2)"
                        # values = (nro_seguimiento,nro_venta,comprador,direccion,referencia,ciudad,cp,vendedor,"Listo Para Retirar",direccion_completa)
                        # cursor.execute(sql,values)
                        # midb.commit()
                        lista_viajes.append(paquete)
        midb.close()
        return render_template("CargaArchivo/data.html",
                                titulo="Carga", 
                                data=lista_viajes, 
                                analizados=cant_viajes, 
                                agregados=flex_agregado, 
                                flex=flex, 
                                repetido=repetido, 
                                auth = session.get("user_auth"))
    else:
        return render_template("CargaArchivo/carga_archivo.html",
                                    titulo="Carga", 
                                    auth = session.get("user_auth"), 
                                    clientes=scriptGeneral.consultar_clientes(database.connect_db()), 
                                    url_post="upload")
