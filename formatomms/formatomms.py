#!/usr/bin/env python
# -*- coding: utf-8 -*- 
# encoding: utf-8

from flask import Blueprint, render_template, request, session,redirect
from datetime import datetime
from auth import auth
from database import database
from informeErrores import informeErrores
import openpyxl
from scriptGeneral import scriptGeneral


formms = Blueprint('formms', __name__, url_prefix='/')

@formms.route("/carga_formms", methods = ["GET","POST"])
@auth.login_required
def subir_exel_formms():
    ahora = (datetime.today())
    if request.method == "POST":
        try:
            midb = database.connect_db()
            cursor = midb.cursor()
            nros_envios = []
            midb = database.verificar_conexion(midb)
            cursor.execute("select Numero_envío from ViajesFlexs")
            for x in cursor:
                nros_envios.append(str(x[0]).lower())
        except:
            print(u"Error al conectar DB!\n")
    # try:
        archivo_xlsx = request.files["upload"]
        libro = openpyxl.load_workbook(archivo_xlsx)
        sheet_obj = libro.active 
        cant_columnas = 14
        contador = 0
    # except:
        # print(u"\nerror en la lectura del archivo\n")
        try:
            # Numero_envío = Order Number
            # nro_venta = Order Number
            # comprador = First Name (Shipping) +	Last Name (Shipping)
            # Telefono = Phone (Billing)
            # Direccion = Address 1&2 (Shipping)
            # Referencia = Customer Note
            # Localidad = City (Shipping)
            # CP = Postcode (Shipping)
            # Cobrar = Order Total Amount
            # Reprogramaciones = 0
            for x in range(cant_columnas):
                contador += 1
                cab = str(sheet_obj.cell(row = 1, column = contador).value)
                if cab.lower() == "fecha":
                    col_fecha = contador
                elif cab.lower() in ("numero de envio", "numero de envío", "nro de envio","order number"):
                    col_numero_envio = contador
                elif cab.lower() in ("cliente","first Name (shipping)","comprador","quien recibe"):
                    col_cliente = contador
                elif cab.lower() in ("direccion","address 1&2 (Shipping)"):
                    col_direccion = contador
                elif cab.lower() in ("localidad","city (shipping)"):
                    col_localidad = contador
                elif cab.lower() in ("vendedor"):
                    col_vendedor = contador
                elif cab.lower() in ("cp","postcode (shipping)","código postal","codigo postal"):
                    col_cp = contador
                elif cab.lower() in ("telefono","phone (billing)"):
                    col_telefono = contador
                elif cab.lower() in ("referencia","customer note","detalle dirección","detalle direccion"):
                    col_referencia = contador
            print("\nasignacion completa\n")
        except Exception as cabezeras:
            informeErrores.reporte(cabezeras,"/carga_formms")
            print("\nError al asignar variables\n")
        n_row = 1
        total = 0
        omitido = 0
        flex_agregado = 0
        blanco = 0
        lista_viajes = []
        actualizados = 0
        cantidad = range(sheet_obj.max_row)
        contadorCantidad = 0
        for fila in cantidad:
            contadorCantidad += 1
            if contadorCantidad > 200:
                break
        if contadorCantidad > 200:
            cantidad = range(0,200)
        for x in cantidad:
            total += 1
            n_row += 1
            if col_numero_envio:
                nro_envio = str(sheet_obj.cell(row = n_row, column = col_numero_envio).value)
            else:
                nro_envio = ""
            if str(nro_envio).lower() in nros_envios or str(nro_envio) == "" or nro_envio == None:
                omitido += 1
            else:
                fecha = str(ahora)[0:10]
                if col_fecha:
                    fecha = str(sheet_obj.cell(row = n_row, column = col_fecha).value)
                if col_cliente:
                    cliente = str(sheet_obj.cell(row = n_row, column = col_cliente).value)
                else:
                    cliente = ""
                if col_telefono:    
                    telefono = str(sheet_obj.cell(row = n_row, column = col_telefono).value)
                else:
                    telefono = None
                if col_direccion: 
                    direccion = str(sheet_obj.cell(row = n_row, column = col_direccion).value)
                else:
                    direccion = ""
                if "/" in str(direccion):
                    direccion = str(direccion.split("/"))[0]
                if col_referencia:
                    referencia = str(sheet_obj.cell(row = n_row, column = col_referencia).value)
                else:
                    referencia = ""
                if col_localidad: localidad = str(sheet_obj.cell(row = n_row, column = col_localidad).value)
                if col_cp:
                    cp = str(sheet_obj.cell(row = n_row, column = col_cp).value)
                else:
                    cp = 0
                if cp == "":
                    cp = 0
                if session.get("user_auth") == "Cliente":
                    vendedor = session.get("user_id")
                elif session.get("user_auth") == "Zippin":
                    vendedor = str(sheet_obj.cell(row = n_row, column = col_vendedor).value)
                estado = "Listo para retirar"
                fecha = fecha[0:10].replace("/","-").replace("\\","")
                tipo_envio = "e-commerce"
                direccion_concatenada = str(direccion) + ", " + str(localidad) + ", Buenos Aires Argentina"     
                chars = '.,!"#$%&/()=?¡¿'
                nro_envio = nro_envio.translate(str.maketrans('', '', chars))
                print(nro_envio)
                nros_envios.append(nro_envio.lower())
                if str(estado).lower() != "no vino": 
                    midb = database.verificar_conexion(midb)
                    cursor.execute("insert into ViajesFlexs (Fecha, Numero_envío, comprador, telefono, Direccion, Referencia, Localidad, CP, Vendedor, tipo_envio, estado_envio, Direccion_Completa) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (fecha,nro_envio,cliente,telefono,direccion,referencia,localidad,cp,vendedor,tipo_envio,estado,direccion_concatenada))
                    midb.commit()
                    flex_agregado += 1
                    
                    
                paquete = [fecha,nro_envio,cliente,telefono,direccion,referencia,localidad,cp,vendedor,tipo_envio,"listo para retirar"]
                lista_viajes.append(paquete)
        print(actualizados)
        cabezera = ["Fecha","Numero de envio","Cliente","Numero de venta","Telefono","Direccion","Referencia","Localidad","cp","Vendedor"]
        midb.close()
        return redirect("/")

    else:
        return render_template("CargaArchivo/carga_archivo.html",
                                titulo="Carga", 
                                clientes=scriptGeneral.consultar_clientes(database.connect_db()), 
                                auth = session.get("user_auth"), 
                                url_post="carga_formms")


