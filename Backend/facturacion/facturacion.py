
from flask import (
    Blueprint, g, render_template, request, session,send_file
)
from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font,Alignment
from Backend.auth import auth
from .estrategiaDeFacturacion import *
from Backend.database.database import connect_db
from Backend.scriptGeneral import scriptGeneral

FcGeneral = Blueprint('facturacionGeneral', __name__, url_prefix='/')

class EnvioAFacturar():
    def __init__(self,id,fecha,nEnvio,direccionCompleta,localidad,precio,comprador,cobrar,estado,motivo,valorDeclarado,estadoActual):
        self.id = id
        self.Fecha = fecha
        self.Numero_envío = nEnvio
        self.Direccion = direccionCompleta
        self.Localidad = localidad
        self.Precio_Cliente = precio
        self.comprador = comprador
        self.Cobrar = cobrar
        self.estado_envio = estado
        self.motivo = motivo
        self.valorDeclarado = valorDeclarado
        self.estadoActual = estadoActual

class Facturador:
    def __init__(self, strategy):
        self.strategy = strategy
        
    def facturar_viajes(self, viajes,sobreEscribir):
        return self.strategy.facturar_viajes(viajes,sobreEscribir)

def generarExcelLiquidacion(envios,_desde,_hasta,_cliente):
    viajes = []
    book = Workbook()
    sheet = book.active
    image = Image("static/logo_grande.jpeg")
    sheet.add_image(image, 'A1')
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15


    sheet["D3"].alignment = Alignment(horizontal='right')
    sheet["D4"].alignment = Alignment(horizontal='right')
    sheet["D5"].alignment = Alignment(horizontal='right')
    sheet["D6"].alignment = Alignment(horizontal='right')
    sheet["D3"].value = "Fecha inicial:"
    sheet["E3"].value = _desde
    sheet["D4"].value = "Fecha final:"
    sheet["E4"].value = _hasta
    sheet["D5"].value = "Vendedor:"
    sheet["E5"].value = _cliente
    sheet["D6"].value = "Cantidad:"
    sheet["E6"].value = len(envios)

    sheet["G2"].value = "Servicio:"
    sheet["H2"].value = "=SUM(F10:F3000)"
    sheet["G3"].value = "Seguro sobre valor declarado:"
    sheet["H3"].value = "=SUM(H10:h3000)*I3"
    sheet["I3"].value = "1%"
    sheet["G4"].font = Font(bold=True)
    sheet["G4"].value = "Subtotal:"
    sheet["H4"].font = Font(bold=True)
    sheet["H4"].value = "=SUM(H2:H3)"
    sheet["G5"].value = "Seguro:"
    sheet["H5"].value = "=H6*I5"
    sheet["I5"].value = "3%"
    sheet["G6"].value = "Cobrar:"
    sheet["H6"].value = "=SUM(G10:G3000)"
    sheet["G7"].font = Font(bold=True)
    sheet["G7"].value = "IVA:"
    sheet["H7"].font = Font(bold=True)
    sheet["H7"].value = "=(H2+H3)*0.21"
    sheet["G8"].value = "Total:"
    sheet["H8"].font = Font(bold=True)
    sheet["H8"].fill = PatternFill(fgColor='00FF00', fill_type='solid')

    sheet["H8"].value = "=H4+H7"
    
    sheet["A9"] = "Fecha"
    sheet["B9"] = "Numero de envío"
    sheet["C9"] = "Comprador"
    sheet["D9"] = "Direccion_Completa"
    sheet["E9"] = "Localidad"
    sheet["F9"] = "Precio"
    sheet["G9"] = "Monto a cobrar"
    sheet["H9"] = "Valor declarado"
    sheet["I9"] = "Estado"
    for row in sheet['A9:I9']:
        for cell in row:
            cell.fill = PatternFill(fgColor='FF0000', fill_type='solid')
            cell.font = Font(color='FFFFFF')
    suma = 0
    contador = 9
    sinprecio = 0
    for viajeTupla in envios:
        viaje = list(viajeTupla)
        contador += 1
        fecha = viaje[0]
        nenvio = viaje[1]
        direccionCompleta = viaje[2]
        localidad = viaje[3]
        precio = viaje[4]
        comprador = viaje[5]
        cobrar = viaje[6]
        valorDeclarado = viaje[7]
        estado = viaje[8]
        if estado != "Entregado":
            cobrar = "0"
        if(precio == None):
            sinprecio = sinprecio +1
            precio = "Sin precio"
        else:
            suma = suma + float(precio)
        sheet["A"+str(contador)] = fecha
        sheet["B"+str(contador)] = nenvio
        sheet["C"+str(contador)] = comprador
        sheet["D"+str(contador)] = direccionCompleta
        sheet["E"+str(contador)] = localidad
        sheet["F"+str(contador)] = precio
        if estado == "Entregado":
            sheet["G"+str(contador)] = cobrar
        else:
            sheet["G"+str(contador)] = "0"
        sheet["H"+str(contador)] = valorDeclarado
        sheet["I"+str(contador)] = estado
        viajes.append(viaje)


    # Crear un objeto AutoFilter para el rango de celdas deseado
    auto_filter = sheet.auto_filter
    auto_filter.ref = 'A9:I3000'
    ruta_archivo = f"{_cliente} {_desde} al {_hasta}.xlsx"
    book.save(ruta_archivo)
    return ruta_archivo


@FcGeneral.route("/descargaresumen/<nombreArchivo>")
@auth.login_required
def descargaResumen(nombreArchivo):
    return send_file(nombreArchivo)



@FcGeneral.route("/facturacion/facturar",methods = ["GET","POST"])
@auth.login_required
def facturar():
    if request.method == "POST":
        cliente = request.form.get("cliente")
        desde = request.form.get("desde")
        hasta = request.form.get("hasta")
        sobreEscribir = request.form.get("sobreEscribir")
        if sobreEscribir == "True":
            sobreEscribir = True
        else:
            sobreEscribir = False
        estrategiaDeFacturacion = request.form.get("estrategiaFacturacion")
        sql = f"""select 
            H.id,
            H.Fecha, 
            H.Numero_envío,
            H.Direccion_Completa,
            H.Localidad,
            H.Precio,
            V.comprador,
            V.Cobrar,
            H.estado_envio,
            H.motivo_noenvio,
            V.valordeclarado,
            V.estado_envio as estadoActual
        from 
            historial_estados as H 
        inner join 
            ViajesFlexs as V 
        on 
            V.Numero_envío = H.Numero_envío 
        where 
            vendedor(H.Vendedor) = '{cliente}' and H.Fecha between '{desde}' and '{hasta}' and H.estado_envio != "Lista para Devolver";"""
        midb = connect_db()
        cursor = midb.cursor()
        cursor.execute(sql)
        viajes = []
        for x in cursor.fetchall():
            viajes.append(EnvioAFacturar(x[0],x[1],x[2],x[3],x[4],x[5],x[6],x[7],x[8],x[9],x[10],x[11]))
        estrategia = EnCaminoStrategy()
        if estrategiaDeFacturacion == "strategyEnCamino":
            estrategia = EnCaminoStrategy()
        elif estrategiaDeFacturacion == "strategyPorDireccion":
            estrategia = EnCaminoUnicoStrategy()
        elif estrategiaDeFacturacion == "strategyEntregado":
            estrategia = EntregadoStrategy()      
        elif estrategiaDeFacturacion == "strategyPorVisita":
            estrategia = PorVisitaStrategy()    
        facturador = Facturador(estrategia)
        total_viajes,viajesEnCamino = facturador.facturar_viajes(viajes,sobreEscribir)
        cabeceras = ["Fecha","Numero de envío","Direccion Completa","Localidad","Precio","Comprador"]
        ruta = generarExcelLiquidacion(viajesEnCamino,desde,hasta,cliente)
        # return send_file(ruta,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, attachment_filename=ruta)
        return render_template("facturacion/tabla_viajes.html",
                            cliente=cliente,
                            desde=desde,
                            hasta=hasta,
                            titulo="Facturacion", 
                            cabeceras = cabeceras,
                            tipo_facturacion="flex", 
                            viajes=viajesEnCamino, 
                            ruta_archivo = ruta,
                            total=f"${total_viajes} y {0} viajes sin precio", 
                            clientes = scriptGeneral.consultar_clientes(midb), 
                            auth = session.get("user_auth")
                            )
    else:
        return render_template("facturacion/tabla_viajes.html",
                            titulo="Facturacion", 
                            desde=str(datetime.now()-timedelta(days=15))[0:10],
                            hasta=str(datetime.now())[0:10],
                            clientes=scriptGeneral.consultar_clientes(connect_db()), 
                            tipo_facturacion="flex", 
                            auth = session.get("user_auth"))
